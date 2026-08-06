from __future__ import annotations

import re
import subprocess
import tempfile
import threading
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET
import xml.parsers.expat as expat

W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
SS_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

# ZIP 炸弹防护：解压后单个文件大小上限（100 MB）
_MAX_UNCOMPRESSED_SIZE = 100 * 1024 * 1024


def _safe_fromstring(xml_bytes: bytes) -> ET.Element:
    """防 XML 实体膨胀（billion laughs）：拒绝任何 DTD 实体声明。"""
    def forbid_entity(*args):
        raise ValueError("XML 实体声明被禁止，防止实体膨胀攻击")

    parser = expat.ParserCreate()
    parser.EntityDeclHandler = forbid_entity
    try:
        parser.Parse(xml_bytes, True)
    except expat.ExpatError:
        # 如果 expat 报语法错误但没触发 entity handler，仍可继续
        pass
    # expat 扫过无实体声明，交给 ElementTree 解析
    return ET.fromstring(xml_bytes)


def _safe_zip_read(zf: zipfile.ZipFile, name: str) -> bytes:
    """防 ZIP 炸弹：按 header 声明的解压大小拦截，再按实际读取量二次校验。"""
    info = zf.getinfo(name)
    if info.file_size > _MAX_UNCOMPRESSED_SIZE:
        raise ValueError(
            f"压缩包内文件过大（{info.file_size} 字节），超过 "
            f"{_MAX_UNCOMPRESSED_SIZE} 字节上限"
        )
    with zf.open(name) as fh:
        data = fh.read(_MAX_UNCOMPRESSED_SIZE + 1)
    if len(data) > _MAX_UNCOMPRESSED_SIZE:
        raise ValueError("压缩包内文件解压后超过大小上限")
    return data


def _read_text_auto(path: Path) -> str:
    """按常见中文编码尝试解码，避免 GBK 文本被当 UTF-8 丢掉。"""
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk", "cp936"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def extract_file(path: Path) -> str:
    """兼容旧调用：只返回提取文本。"""
    return extract_file_with_status(path)[0]


def extract_file_with_status(path: Path) -> tuple[str, str, str]:
    """提取文本并返回状态标识。

    返回 (text, status, error)，status 取值：
    - success: 提取到有效文本
    - no_text: 无有效文本（图片未 OCR、暂不支持类型、空文本等）
    - failed:  提取过程异常
    """
    ext = path.suffix.lower()
    try:
        if ext == ".docx":
            text = _extract_docx(path)
        elif ext in {".xlsx", ".xlsm"}:
            text = _extract_xlsx(path)
        elif ext == ".pdf":
            text = _extract_pdf(path)
        elif ext == ".doc":
            text = _extract_doc(path)
        elif ext in {".txt", ".md", ".csv"}:
            text = _read_text_auto(path)
        elif ext in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}:
            return _ocr_image_with_status(path)
        else:
            return f"[暂不支持的文件类型: {path.name}]", "no_text", f"暂不支持的文件类型: {ext}"
    except Exception as exc:  # noqa: BLE001
        return f"[解析失败 {path.name}: {exc}]", "failed", str(exc)

    text = (text or "").strip()
    if not text or text.startswith("["):
        if text.startswith("[解析失败"):
            return text, "failed", text
        return text, "no_text", text or "未能从文件中提取到文本"
    return text, "success", ""


def _extract_docx(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        xml = _safe_zip_read(zf, "word/document.xml")
    root = _safe_fromstring(xml)
    blocks: list[str] = []

    body = root.find(f"{W_NS}body")
    if body is None:
        return ""

    for child in list(body):
        tag = child.tag
        if tag == f"{W_NS}p":
            text = _para_text(child)
            if text:
                blocks.append(text)
        elif tag == f"{W_NS}tbl":
            blocks.append(_table_markdown(child))

    return "\n".join(blocks).strip()


def _para_text(p: ET.Element) -> str:
    parts = [t.text for t in p.iter(f"{W_NS}t") if t.text]
    return "".join(parts).strip()


def _table_markdown(tbl: ET.Element) -> str:
    rows: list[list[str]] = []
    for tr in tbl.findall(f"{W_NS}tr"):
        cells: list[str] = []
        for tc in tr.findall(f"{W_NS}tc"):
            texts = [t.text for t in tc.iter(f"{W_NS}t") if t.text]
            cells.append(re.sub(r"\s+", " ", "".join(texts)).strip())
        if any(cells):
            rows.append(cells)
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    norm = [r + [""] * (width - len(r)) for r in rows]
    lines = ["| " + " | ".join(norm[0]) + " |"]
    lines.append("| " + " | ".join(["---"] * width) + " |")
    for r in norm[1:]:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


def _extract_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _extract_xlsx_raw(path)

    wb = load_workbook(path, data_only=True, read_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        parts.append(f"## 工作表: {sheet.title}")
        rows_out: list[str] = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            vals = ["" if c is None else str(c).replace("\n", " ").strip() for c in row]
            if not any(vals):
                continue
            rows_out.append(" | ".join(vals))
            if i > 500:
                rows_out.append("... (已截断)")
                break
        parts.append("\n".join(rows_out))
    wb.close()
    return "\n\n".join(parts).strip()


def _extract_xlsx_raw(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = _safe_fromstring(_safe_zip_read(zf, "xl/sharedStrings.xml"))
            for si in root.findall(f"{SS_NS}si"):
                texts = [t.text or "" for t in si.iter(f"{SS_NS}t")]
                shared.append("".join(texts))
        sheets = sorted(n for n in zf.namelist() if n.startswith("xl/worksheets/sheet"))
        parts: list[str] = []
        for sh in sheets:
            root = _safe_fromstring(_safe_zip_read(zf, sh))
            lines: list[str] = []
            for row in root.findall(f".//{SS_NS}row"):
                vals: list[str] = []
                for c in row.findall(f"{SS_NS}c"):
                    t = c.get("t")
                    v = c.find(f"{SS_NS}v")
                    if v is None or v.text is None:
                        vals.append("")
                        continue
                    val = v.text
                    if t == "s" and val.isdigit():
                        idx = int(val)
                        val = shared[idx] if idx < len(shared) else val
                    vals.append(val.replace("\n", " "))
                if any(vals):
                    lines.append(" | ".join(vals))
            parts.append("\n".join(lines))
        return "\n\n".join(parts).strip()


def _ocr_image_with_status(path: Path) -> tuple[str, str, str]:
    """单张图片走 RapidOCR 识别文字。"""
    try:
        engine = _get_ocr_engine()
    except ImportError:
        return (
            f"[未安装 rapidocr-onnxruntime，无法 OCR 图片: {path.name}]",
            "no_text",
            "未安装 OCR 依赖",
        )
    try:
        result, _ = _run_ocr(engine, str(path))
        if not result:
            return "", "no_text", "OCR 未识别到文字"
        return "\n".join(item[1] for item in result).strip(), "success", ""
    except Exception as exc:  # noqa: BLE001
        return f"[图片 OCR 失败 {path.name}: {exc}]", "failed", str(exc)


def _extract_pdf(path: Path) -> str:
    try:
        import fitz  # pymupdf
    except ImportError:
        return f"[未安装 pymupdf，无法解析 PDF: {path.name}]"

    doc = fitz.open(path)
    texts: list[str] = []
    max_pages = min(doc.page_count, 40)
    for i in range(max_pages):
        page = doc.load_page(i)
        texts.append(page.get_text("text"))
    doc.close()
    text = "\n".join(texts).strip()
    if text:
        return text
    return _ocr_pdf(path, max_pages)


_ocr_engine: object | None = None
# 初始化锁与推理锁分开：初始化在推理锁外，避免同一把非可重入锁自死锁
_ocr_lock = threading.Lock()
_ocr_infer_lock = threading.Lock()


def _run_ocr(engine, target: str):
    """串行执行 OCR 推理。RapidOCR/onnxruntime session 未声明线程安全，
    而调用方在 threadpool 中并发进入，这里加锁避免竞态。"""
    with _ocr_infer_lock:
        return engine(target)


def _get_ocr_engine():
    """懒加载 RapidOCR 引擎（首次初始化较慢，之后复用）。线程安全。"""
    global _ocr_engine
    with _ocr_lock:
        if _ocr_engine is None:
            from rapidocr_onnxruntime import RapidOCR

            _ocr_engine = RapidOCR()
        return _ocr_engine


def _ocr_pdf(path: Path, max_pages: int) -> str:
    """无文字层的扫描件 PDF：渲染成图片后用 RapidOCR 识别中文。"""
    try:
        engine = _get_ocr_engine()
    except ImportError:
        return f"[PDF 无文字层，且未安装 rapidocr-onnxruntime: {path.name}]"

    import fitz  # pymupdf

    doc = fitz.open(path)
    parts: list[str] = []
    try:
        with tempfile.TemporaryDirectory(prefix="detection_ocr_") as tmp:
            for i in range(max_pages):
                page = doc.load_page(i)
                pix = page.get_pixmap(dpi=200)
                png = Path(tmp) / f"page_{i}.png"
                pix.save(png)
                result, _ = _run_ocr(engine, str(png))
                if result:
                    lines = [item[1] for item in result]
                    parts.append("\n".join(lines))
    finally:
        doc.close()
    return "\n".join(parts).strip()


def _extract_doc(path: Path) -> str:
    # macOS
    try:
        result = subprocess.run(
            ["textutil", "-convert", "txt", "-stdout", str(path)],
            capture_output=True,
            text=True,
            timeout=60,
            check=False,
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout.strip()
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass

    # LibreOffice
    try:
        out_dir = path.parent
        result = subprocess.run(
            [
                "soffice",
                "--headless",
                "--convert-to",
                "txt:Text",
                "--outdir",
                str(out_dir),
                str(path),
            ],
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
        )
        txt_path = out_dir / f"{path.stem}.txt"
        if txt_path.exists():
            return _read_text_auto(txt_path).strip()
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass

    return f"[无法解析 .doc，请另存为 .docx 后重新上传: {path.name}]"


def combine_project_texts(file_texts: list[tuple[str, str]], max_chars: int) -> str:
    chunks: list[str] = []
    for name, text in file_texts:
        chunks.append(f"===== 文件: {name} =====\n{text}")
    combined = "\n\n".join(chunks)
    if len(combined) > max_chars:
        combined = combined[:max_chars] + "\n\n...[内容过长已截断]"
    return combined
